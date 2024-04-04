import os
import time

import openpyxl

from elf_py_utils import logger_util

logger = logger_util.set_logging('ExcelUtil')


def read_excel(title_row: int, file_name: str) -> list[dict]:
    """
    读取excel文件，返回一列表，列表中元素为字典

    :param title_row: 列名所在的行索引
    :param file_name: .xlsx文件地址
    :return: excel文件内容，列表中每个元素即为表格中的一行数据
    """
    start_time = time.time()
    excel_content = []
    # 加载excel
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    # 获取excel表格的列名
    column_name = []
    for y in range(1, ws.max_column + 1):
        column_name.append(ws.cell(title_row, y).value)
    for x in range(title_row + 1, ws.max_row + 1):
        record = {}
        for y in range(1, ws.max_column + 1):
            record[column_name[y - 1]] = ws.cell(x, y).value
        excel_content.append(record)
        if x % 500 == 0:
            logger.info('[读取excel]总行数：%s，当前行数：%s，进度：%s%s，当前耗时：%s秒',
                        ws.max_row, x, round(x / ws.max_row * 100, 2), '%', round(time.time() - start_time))
    return excel_content


def read_csv(title_row: int, file_name: str) -> list[dict]:
    """
    读取csv文件，返回一列表，列表中元素为字典

    :param title_row:
    :param file_name:
    :return: list[dict]
    """
    logger.info('正在读取csv文件')
    start_time = time.time()
    csv_content = []
    # 加载csv文件
    csv_file = open(file_name, 'r', encoding='utf-8')
    # 获取csv文件的列名
    column_name = csv_file.readline().replace('\n', '').split(',')
    n = 1
    while True:
        n += 1
        row = csv_file.readline()
        if row:
            record = {}
            row_list = row.replace('\n', '').split(',')
            if len(row_list) == len(column_name):
                if n % 500000 == 0:
                    logger.info('当前行数为：%d', n)
                for i in range(len(row_list)):
                    record[column_name[i]] = row_list[i]
                csv_content.append(record)
            else:
                logger.info('csv文件的第%d行数据格式有误', n)
                logger.info('其内容为：%s', row)
                break
        else:
            logger.info('csv文件读取完毕')
            break
    end_time = time.time()
    logger.info('处理csv文件耗时为：%.3f秒', (end_time - start_time))
    return csv_content


def read_folder(folder_path: str, suffix: str = '') -> list:
    """读取文件夹，获取文件夹中所有文件名列表并返回

    * 返回结果仅含有.suffix文件

    :param folder_path: 文件夹路径
    :param suffix: 文件后缀
    :return: list
    """
    files = os.listdir(folder_path)
    i = 0
    while i < len(files):
        if suffix != '':
            if suffix != files[i][-len(suffix):]:
                files.pop(i)
            else:
                i += 1
        else:
            i += 1
    return files
