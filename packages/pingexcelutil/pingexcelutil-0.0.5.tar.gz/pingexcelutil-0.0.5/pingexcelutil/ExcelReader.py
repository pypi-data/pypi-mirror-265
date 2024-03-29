import pandas as pd
from openpyxl import load_workbook


class ExcelReader():
    def __init__(self, excel_file):
        self.excel_file = excel_file
        pass

    def convert_list_to_pandas(self, list_rows, list_header=None):
        if list_header:
            df = pd.DataFrame(list_rows, list_header)
        else:
            df = pd.DataFrame(list_rows)
        return df

    def select_sheet(self, by, sheet_name):
        select_by = None
        if not by:
            select_by = "sheet_name"
        else:
            select_by = "sheet_order"

        return select_by

    def read_from_ranges(self, sheet_name, col_range):
        excel_file = self.excel_file

        try:
            sheet_order = int(sheet_name)
            wb = load_workbook(filename=excel_file, read_only=True, data_only=True)
            ws = wb[sheet_order]
        except Exception as e:
            print(str(e))
            return False

        list_rows = []

        for row in ws.iter_rows(col_range.format(ws.min_row, ws.max_row)):
            list_cols = []
            for cell in row:
                list_cols.append(str(cell.value))
            list_rows.append(list_cols)

        return list_rows

    def read_excel_from_whole_columns2(self, excel_file_name, sheet_order, col_range):
        wb = load_workbook(filename=excel_file_name, read_only=True, data_only=True)
        ws = wb.worksheets[sheet_order]
        rows = []

        for row in ws.iter_rows(col_range.format(ws.min_row, ws.max_row)):
            cols = []
            for cell in row:
                cols.append(cell.value)
            rows.append(cols)
        df_excel = pd.DataFrame(rows)
        return df_excel

    def read_excel_from_range(self, excel_file_name, sheet_name, range_from, range_to):
        wb = load_workbook(filename=excel_file_name, read_only=True, data_only=True)
        ws = wb[sheet_name]

        rows = []
        data_range = ws[range_from:range_to]

        for i, row in enumerate(data_range):
            cols = []
            for j, cell in enumerate(row):
                cols.append(cell.value)
                if j == len(row) - 1:
                    cols.append(sheet_name)
                    cols.append(i)
                    cols.append(excel_file_name)
            rows.append(cols)
        df_excel = pd.DataFrame(rows)
        return df_excel
