from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from ficamp.datastructures import Tx
from ficamp.parsers.protocols import Parser


class AccountBBVAParser(Parser):
    """Parser for BBVA bank account extract"""

    def load(self, filename: Path | None = None):
        # TODO: rearrange this.

        # filename = Path("../data/enero-febrero-bbva-cuenta.xlsx")
        # filename = os.path.join(
        #     os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        #     "data/enero-febrero-bbva-cuenta.xlsx",
        # )

        wb = load_workbook(filename)
        sheet = wb.active
        start_row = 6
        start_column = 2

        return [
            row
            for row in sheet.iter_rows(
                min_row=start_row, min_col=start_column, values_only=True
            )
        ]

    def parse(self) -> list[Tx]:
        rows = self.load()

        return [
            self.row_processor(row)
            for row in rows
            if self.row_processor(row) is not None
        ]

    def row_processor(self, row):
        # Skip Credit Card charge in Account
        if "targeta" in row[2] or "tarjeta" in row[2]:
            return None

        if row[5] == "EUR":
            currency = "€"

        concept = f"{row[2]} || {row[3]}"

        return Tx(
            date=row[0],
            amount=Decimal(str(row[4])),
            currency=currency,
            concept=concept,
            category=None,
            metadata={"more_details": row[8], "origin": "BBVA Account"},
            tags=[],
        )


class CreditCardBBVAParser(Parser):
    """Parser for BBVA Credit Card Extract"""

    def load(self, filename: Path | None = None):
        # TODO: rearrange this
        # filename = Path("../data/enero-febrero-bbva-cuenta.xlsx")
        # filename = os.path.join(
        #     os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        #     "data/enero-febrero-bbva-targeta.xlsx",
        # )

        wb = load_workbook(filename)
        sheet = wb.active
        start_row = 6
        start_column = 2

        return [
            row
            for row in sheet.iter_rows(
                min_row=start_row, min_col=start_column, values_only=True
            )
        ]

    def parse(self) -> list[Tx]:
        rows = self.load()

        return [
            self.row_processor(row)
            for row in rows
            if self.row_processor(row) is not None
        ]

    def row_processor(self, row):
        # Skip Recharging the Credit Card, as it's an useless operation from user's POV.
        if row[3] > 0:
            return None

        currency = "€"

        return Tx(
            date=datetime.strptime(row[0], "%d/%m/%Y"),
            amount=Decimal(str(row[3])),
            currency=currency,
            concept=row[2],
            category=None,
            metadata={"origin": "BBVA Credit Card"},
            tags=[],
        )


if __name__ == "__main__":
    bbva = AccountBBVAParser()
    txs = bbva.parse()

    txs += CreditCardBBVAParser().parse()

    for tx in txs:
        print(tx)
