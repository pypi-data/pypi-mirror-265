import os
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from ficamp.datastructures import Tx
from ficamp.parsers.protocols import Parser


class AccountBSabadellParser(Parser):
    """Parser for BBVA bank account extract"""

    def load(self, filename: Path | None = None):
        # TODO: rearrange this.

        # filename = Path("../data/enero-febrero-bbva-cuenta.xlsx")
        # filename = os.path.join(
        #     os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        #     "data/enero-febrero-bsabadell-cuenta.xlsx",
        # )

        wb = load_workbook(filename)
        sheet = wb.active
        start_row = 10
        start_column = 1

        return [
            row
            for row in sheet.iter_rows(
                min_row=start_row, min_col=start_column, values_only=True
            )
        ]

    def parse(self) -> list[Tx]:
        rows = self.load()

        return [
            self.row_processor(row)
            for row in rows
            if self.row_processor(row) is not None
        ]

    def row_processor(self, row):
        # Skip Credit Card charge in Account
        if "TARJETA CREDITO" in row[2]:
            return None

        currency = "€"

        concept = self.concept_builder(row)

        return Tx(
            date=datetime.strptime(row[0], "%d/%m/%Y"),
            amount=Decimal(str(row[3])),
            currency=currency,
            concept=concept,
            category=None,
            metadata={"origin": "BSABADELL Account"},
            tags=[],
        )

    def concept_builder(self, row):
        """
        There are some rows with more details than others.
        This function builds the concept with all the details available.
        """
        concept = row[1]
        if row[5]:
            concept += f" || {row[5]}"
        if row[6]:
            concept += f" || {row[6]}"

        return concept


class CreditCardBSabadellParser(Parser):
    """Parser for Banc Sabadell Credit Card Extract"""

    def load(self, filename: Path | None = None):
        # TODO: rearrange this
        # filename = Path("../data/enero-febrero-bbva-cuenta.xlsx")
        # filename = os.path.join(
        #     os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        #     "data/enero-febrero-bsabadell-targeta.xlsx",
        # )

        wb = load_workbook(filename)
        sheet = wb.active
        start_row = 12
        start_column = 1

        rows = []
        for row in sheet.iter_rows(
            min_row=start_row, min_col=start_column, values_only=True
        ):
            if row[0] is None:
                # Mixed content for xlsx file, so we need to break when we reach the end of the table
                break

            rows.append(row)
        return rows

    def parse(self) -> list[Tx]:
        rows = self.load()
        return [
            self.row_processor(row)
            for row in rows
            if self.row_processor(row) is not None
        ]

    def row_processor(self, row):
        currency = "€"

        return Tx(
            date=datetime.strptime(f"{row[0]}/{datetime.now().year}", "%d/%m/%Y"),
            amount=Decimal(str(row[4]).replace(",", ".")),
            currency=currency,
            concept=row[1],
            category=None,
            metadata={"origin": "BSABADELL Credit Card", "location": row[2]},
            tags=[],
        )


if __name__ == "__main__":
    bsabadell = AccountBSabadellParser()
    txs = bsabadell.parse()

    txs += CreditCardBSabadellParser().parse()

    for tx in txs:
        print(tx.concept)
