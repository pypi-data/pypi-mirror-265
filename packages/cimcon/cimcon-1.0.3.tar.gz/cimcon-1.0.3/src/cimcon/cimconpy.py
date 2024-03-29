#-----------------------------------------------Packages-------------------------------------
import io
import os
import shap
import sys
import pickle
import numpy as np
import pandas as pd
import seaborn as sns
from datetime import datetime
import matplotlib.pyplot as plt
from PIL import Image as PILImage
from openpyxl.drawing.image import Image
from mapie.regression import MapieRegressor
from openpyxl import load_workbook, Workbook
from fairlearn.metrics import false_positive_rate, false_negative_rate, selection_rate, count, MetricFrame
from sklearn.metrics import precision_score, accuracy_score, confusion_matrix, roc_curve, auc, precision_recall_curve

import warnings
warnings.simplefilter(action='ignore')

class CIMCONTest:
    def __init__(self):
        pass
    
    def _save_data_with_images(self, data, plt_figures, filename, model_name):
        """Save data with images to a specified filepath."""
        try:
            # Try loading existing workbook
            wb = load_workbook(filename=f"{filename}.xlsx")
        except FileNotFoundError:
            # If workbook doesn't exist, create new
            wb = Workbook()
            # Remove the default sheet created automatically
            wb.remove(wb.active)
    
        # Create a new worksheet
        ws = wb.create_sheet()
    
        # Add DataFrame or Series to the worksheet
        if isinstance(data, pd.DataFrame):
            ws.append(data.columns.tolist())  # Add column names
            for _, row in data.iterrows():
                ws.append(row.tolist())
        elif isinstance(data, pd.Series):
            ws.append(data.index.tolist())  # Add index as column name for Series
            ws.append(data.tolist())
    
        # Calculate the starting position for the image
        start_row = 2  # Start from the row after the data
        start_col = len(data.columns) + 1  # Start from the column after the data
    
        # Convert matplotlib figures to images and concatenate them
        images = []
        for fig in plt_figures:
            buf = io.BytesIO()
            fig.savefig(buf, format='png')
            buf.seek(0)
            pil_image = PILImage.open(buf)
            images.append(pil_image)
    
        # Concatenate images horizontally
        total_width = sum(image.size[0] for image in images)
        max_height = max(image.size[1] for image in images)
        concatenated_image = PILImage.new('RGB', (total_width, max_height))
        x_offset = 0
        for image in images:
            concatenated_image.paste(image, (x_offset, 0))
            x_offset += image.size[0]
    
        # Convert the concatenated image to an openpyxl Image
        img_io = io.BytesIO()
        concatenated_image.save(img_io, format='PNG')
        img_io.seek(0)
        img = Image(img_io)
    
        # Add the image to the worksheet
        ws.add_image(img, f'{chr(ord("A") + start_col)}{start_row}')  # Adjust the cell as needed
    
        # Rename the worksheet with the model name
        ws.title = model_name
    
        # Save the workbook
        wb.save(f"{filename}.xlsx")

    def fair_metrics(self, X_test, y_test, y_predicted, sensitive_feature):
        """Conduct fairness tests with respective features."""
        try:
            # Read the sensitive feature
            protected_class = X_test[sensitive_feature]

            # Define fairness metrics to compute
            metrics = {
                "accuracy": accuracy_score,
                "precision": precision_score,
                "false positive rate": false_positive_rate,
                "false negative rate": false_negative_rate,
                "selection rate": selection_rate,
                "count": count,
            }

            # Compute fairness metrics using MetricFrame
            metric_frame = MetricFrame(
                metrics=metrics, y_true=y_test, y_pred=y_predicted, sensitive_features=protected_class
            )
            # Create a plot to visualize all fairness metrics
            fig, ax = plt.subplots(figsize=(12, 13))
            image_temp = metric_frame.by_group.plot.bar(
                subplots=True,
                layout=[3, 3],
                legend=False,
                figsize=[12, 8],
                colormap='RdBu',
                title="Show all metrics",
                ax=ax
            )
            # Extract computed metrics and sensitive levels
            acc = metric_frame.by_group["accuracy"]
            precise = metric_frame.by_group["precision"]
            fpr = metric_frame.by_group["false positive rate"]
            fnr = metric_frame.by_group["false negative rate"]
            selectionrate = metric_frame.by_group["selection rate"]
            counts = metric_frame.by_group["count"]
            sensitive = metric_frame.sensitive_levels

            # Prepare output data for further analysis or storage
            output_data = pd.DataFrame({'y_test_value': y_test, 'predicted_value': y_predicted,
                                        'class label': np.hstack(([0, 1], [None] * (len(y_test) - 2))),
                                        'accuracy': np.hstack((acc, [None] * (len(y_test) - len(acc)))),
                                        'precision': np.hstack((precise, [None] * (len(y_test) - len(precise)))),
                                        'false positive rate': np.hstack((fpr, [None] * (len(y_test) - len(fpr)))),
                                        'false negative rate': np.hstack((fnr, [None] * (len(y_test) - len(fnr)))),
                                        'selection rate': np.hstack((selectionrate, [None] * (len(y_test) - len(selectionrate)))),
                                        'count': np.hstack((counts, [None] * (len(y_test) - len(counts)))),
                                        'sensitive_levels': sensitive}, index=y_test.index)
            # Save output data and plot
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S%f")

            path = input('Enter the folder path where you want to save the file for eg. C:\\Users\\Desktop\\lib')
            filepath = os.path.join(path, f'CIMCON FairnessTest{formatted_datetime}')

            self._save_data_with_images(output_data, [fig], filepath, 'Fairness Test')

        except Exception as e:
            # Catch any exceptions and provide a descriptive error message
            print(f'Check the Input data,make sure your target value in 0 1 form Unsupported Multi-class classification/ Continuous data: {e}')

#-------------------------------------Interpret--------------------------------------------
    def InterpretabilityBinaryClassTreeTest(self, model, X_test, FeatureNames):
        '''This method calculates the Interpretability of Model'''
        try: 
            # Initialize SHAP Tree Explainer
            explainer = shap.TreeExplainer(model)
            # Calculate SHAP values
            shap_values = explainer.shap_values(X_test)

            # Plot summary SHAP plot
            fig1, ax1 = plt.subplots()
            shap.summary_plot(shap_values[0], X_test, feature_names=FeatureNames, show=False)

            # Plot bar plot of SHAP values
            fig2, ax2 = plt.subplots()
            shap.summary_plot(shap_values, X_test, feature_names=FeatureNames, show=False, plot_type="bar")

            # Create output dataframes
            output_data1 = pd.DataFrame()
            output_data2 = pd.DataFrame()
            for i in range(len(FeatureNames)):
                output_data1[FeatureNames[i]] = shap_values[0][i]
                output_data2[FeatureNames[i] + '_0'] = shap_values[0][i]
                output_data2[FeatureNames[i] + '_1'] = shap_values[1][i]

            # Get current date and time
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S%f")

            # Get file path from user input
            path = input('Enter the folder path where you want to save the file for eg. C:\\Users\\Desktop\\lib')
            filepath = os.path.join(path, '\CIMCON Reg InterpretabilityTest' + formatted_datetime)
            
            # Save output data with images
            self._save_data_with_images(output_data1, [fig1], filepath, model.__class__.__name__)
            self._save_data_with_images(output_data2, [fig2], filepath, model.__class__.__name__)
       
        except Exception as e:
            # Print error message if unsupported input model
            print(f'Unsupported input model {e}')
            
    def InterpretabilityBinaryClassNonTreeTest(self, model, X_train, X_test, FeatureNames):
        '''This method calculates the Interpretability of Non-Tree Models (KNN, NB)'''
        try:
            # Initialize SHAP Explainer for Non-Tree Models
            explainer = shap.Explainer(model.predict, X_train)
            # Calculate SHAP values
            shap_values = explainer.shap_values(X_test) 

            # Plot summary SHAP plot
            fig1, ax1 = plt.subplots()
            shap.summary_plot(shap_values, X_test, feature_names=X_test.columns, show=False)

            # Plot bar plot of SHAP values
            fig2, ax2 = plt.subplots()
            shap.summary_plot(shap_values, X_test, feature_names=X_test.columns, show=False, plot_type="bar")

            # Create output dataframe
            output_data = pd.DataFrame()
            for i in range(len(FeatureNames)):
                output_data[FeatureNames[i]] = shap_values[i]

            # Get current date and time
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S%f")

            # Get file path from user input
            path = input('Enter the folder path where you want to save the file for eg. C:\\Users\\Desktop\\lib')
            filepath = os.path.join(path, '\CIMCON NonTree BinaryClass InterpretabilityTest' + formatted_datetime)
            
            # Save output data with images
            self._save_data_with_images(output_data, [fig1], filepath, model.__class__.__name__)
            self._save_data_with_images(output_data, [fig2], filepath, model.__class__.__name__)

        except Exception as e:
            # Print error message if unsupported input model
            print(f'Unsupported input model {e}')


    def InterpretabilityBinaryClassSVCTest(self, model, X_train, X_test, FeatureNames):
        '''This method calculates the Interpretability of Support Vector Classifier (SVC)'''
        try:
            # Initialize SHAP Kernel Explainer for SVC
            explainer = shap.KernelExplainer(model.predict_proba, X_train, link="logit")
            # Calculate SHAP values
            shap_values = explainer.shap_values(X_test)  

            # Plot summary SHAP plot
            fig1, ax1 = plt.subplots()
            shap.summary_plot(shap_values[0], X_test, feature_names=FeatureNames, show=False)

            # Plot bar plot of SHAP values
            fig2, ax2 = plt.subplots()
            shap.summary_plot(shap_values, X_test, feature_names=FeatureNames, show=False, plot_type="bar")

            # Create output dataframes
            output_data1 = pd.DataFrame()
            output_data2 = pd.DataFrame()
            for i in range(len(FeatureNames)):
                output_data1[FeatureNames[i]] = shap_values[0][i]
                output_data2[FeatureNames[i] + '_0'] = shap_values[0][i]
                output_data2[FeatureNames[i] + '_1'] = shap_values[1][i]

            # Get current date and time
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S%f")

            # Get file path from user input
            path = input('Enter the folder path where you want to save the file for eg. C:\\Users\\Desktop\\lib')
            filepath = os.path.join(path, '\CIMCON SVC InterpretabilityTest' + formatted_datetime)

            # Save output data with images
            self._save_data_with_images(output_data1, [fig1], filepath, model.__class__.__name__)
            self._save_data_with_images(output_data2, [fig2], filepath, model.__class__.__name__)

        except Exception as e:
            # Print error message if unsupported input model
            print(f'Unsupported input model {e}')


    def InterpretabilityRegTest(self, model, X_train, X_test, FeatureNames):
        '''This method calculates the Interpretability of Regression Model'''
        try:
            # Initialize SHAP Explainer for Regression Model
            explainer = shap.Explainer(model.predict, X_train)
            # Calculate SHAP values
            shap_values = explainer.shap_values(X_test)

            # Plot summary SHAP plot
            fig1, ax1 = plt.subplots()
            shap.summary_plot(shap_values, X_test, feature_names=FeatureNames, show=False)

            # Plot bar plot of SHAP values
            fig2, ax2 = plt.subplots()
            shap.summary_plot(shap_values, X_test, feature_names=FeatureNames, show=False, plot_type="bar")

            # Create output dataframe
            output_data = pd.DataFrame()
            for i in range(len(FeatureNames)):
                output_data[FeatureNames[i]] = shap_values[i]

            # Get current date and time
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S%f")

            # Get file path from user input
            path = input('Enter the folder path where you want to save the file for eg. C:\\Users\\Desktop\\lib')
            filepath = os.path.join(path, '\CIMCON Reg InterpretabilityTest' + formatted_datetime)

            # Save output data with images
            self._save_data_with_images(output_data, [fig1], filepath, model.__class__.__name__)
            self._save_data_with_images(output_data, [fig2], filepath, model.__class__.__name__)

        except Exception as e:
            # Print error message if unsupported input model
            print(f'Unsupported input model {e}')


    def ValidityReliabilityRegTest(self, model, X_train, y_train, X_test, y_test):
        '''This method evaluates the performance of the Regression model'''
        try:
            # Initialize MapieRegressor
            mapie_reg = MapieRegressor(estimator=model)
            # Fit Mapie on the training data
            mapie_reg.fit(X_train, y_train)
            # Predict and get prediction intervals on test data
            y_pred, y_pred_intervals = mapie_reg.predict(X_test, alpha=0.05)

            # Scatter Plot of Predictions vs Actual Values
            fig1, ax1 = plt.subplots()
            ax1.scatter(y_test, y_pred, alpha=0.5)
            ax1.plot([y_test.min(), y_test.max()], [y_test.min(), y_test.max()], 'k--', lw=2)
            ax1.set_xlabel('Actual')
            ax1.set_ylabel('Predicted')
            ax1.set_title('Predicted vs Actual Values');

            # Prediction Interval Plot
            lower_bounds, upper_bounds = y_pred_intervals[:, 0, 0], y_pred_intervals[:, 1, 0]
            fig2, ax2 = plt.subplots()
            ax2.fill_between(range(len(y_test)), lower_bounds, upper_bounds, color='skyblue', alpha=0.5)
            ax2.plot(y_test, 'o')
            ax2.plot(y_pred, 'x', color='red')
            ax2.set_title('Prediction Intervals')
            ax2.set_ylabel('Prediction')
            ax2.set_xlabel('Sample');

            # Residual Plot
            fig3, ax3 = plt.subplots()
            residuals = y_test - y_pred
            ax3.scatter(y_pred, residuals)
            ax3.hlines(y=0, xmin=y_pred.min(), xmax=y_pred.max(), colors='red')
            ax3.set_xlabel('Predicted')
            ax3.set_ylabel('Residuals')
            ax3.set_title('Residual Plot');

            # Create output dataframes
            output_data1 = pd.DataFrame({'y_test_value': y_test, 'predicted_value': y_pred, 'y_test.min': y_test.min(),
                                         'y_test.max': y_test.max()}, index=y_test.index)
            output_data2 = pd.DataFrame({'y_test_value': y_test, 'predicted_value': y_pred, 'lower bounds': lower_bounds,
                                         'upper bound': upper_bounds}, index=y_test.index)
            output_data3 = pd.DataFrame({'y_test_value': y_test, 'predicted_value': y_pred, 'Residuals': residuals},
                                        index=y_test.index)

            # Get current date and time
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S%f")

            # Get file path from user input
            path = input('Enter the folder path where you want to save the file for eg. C:\\Users\\Desktop\\lib')
            filepath = os.path.join(path, '\CIMCON Validity Reliability RegTest' + formatted_datetime)

            # Save output data with images
            self._save_data_with_images(output_data1, [fig1], filepath, model.__class__.__name__)
            self._save_data_with_images(output_data2, [fig2], filepath, model.__class__.__name__)
            self._save_data_with_images(output_data3, [fig3], filepath, model.__class__.__name__)
        except Exception as e:
            # Print error message
            print(f'Error: {e}')


    def ValidityReliabilityBinaryclassTest(self, model, X_test, y_test, y_pred):
        '''This method evaluates the performance of the BinaryClass Classification model'''
        try:
            # Confusion Matrix
            fig1, ax1 = plt.subplots()
            cm = confusion_matrix(y_test, y_pred)
            s = sns.heatmap(cm, annot=True, cmap="RdBu", fmt='g', ax=ax1)
            ax1.set_title('Confusion Matrix')
            ax1.set_ylabel('Actual')
            ax1.set_xlabel('Predicted');

            # ROC Curve
            fig2, ax2 = plt.subplots()
            fpr, tpr, thresholds = roc_curve(y_test, model.predict_proba(X_test)[:, 1])
            roc_auc = auc(fpr, tpr)

            ax2.plot(fpr, tpr, lw=2, label=f'ROC curve (AUC = {roc_auc:.2f})')
            ax2.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--', label='y = x^2')
            ax2.set_xlabel('False Positive Rate', fontsize=12)
            ax2.set_ylabel('True Positive Rate', fontsize=12)
            ax2.set_title('ROC Curve', fontsize=16)
            ax2.legend(loc="lower right")
            ax2.grid(which='major', linestyle='--', linewidth='0.5', color='grey');

            # Precision-Recall Curve
            y_score = model.predict_proba(X_test)[:, 1]
            precision, recall, _ = precision_recall_curve(y_test, y_score)

            fig3, ax3 = plt.subplots()
            ax3.plot(recall, precision, lw=2, label='Precision-Recall curve')
            ax3.set_xlabel('Recall')
            ax3.set_ylabel('Precision')
            ax3.set_title('Precision-Recall Curve')
            ax3.legend(loc="lower left");

            # Create output dataframes
            output_data1 = pd.DataFrame({'y_test_value': y_test, 'predicted_value': y_pred, 'True Negative': cm[0][0], 
                                         'False Positive': cm[0][1], 'False Negative': cm[1][0], 'True Positive': cm[1][1]},
                                        index=y_test.index)

            output_data2 = pd.DataFrame({'y_test_value': y_test, 'False Positive Rate': np.hstack((fpr, [None] * (len(y_test) - fpr.shape[0]))),
                                         'True positive rate': np.hstack((tpr, [None] * (len(y_test) - tpr.shape[0]))),
                                         'Thresholds': np.hstack((thresholds, [None] * (len(y_test) - thresholds.shape[0]))),
                                         'roc_auc': roc_auc}, index=y_test.index)

            output_data3 = pd.DataFrame({'y_test_value': y_test, 'y_score': y_score, 'Precision': np.hstack((precision, [None] * (len(y_test) - precision.shape[0]))), 
                                         'Recall': np.hstack((recall, [None] * (len(y_test) - recall.shape[0])))}, index=y_test.index)

            # Get current date and time
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S%f")

            # Get file path from user input
            path = input('Enter the folder path where you want to save the file for eg. C:\\Users\\Desktop\\lib')
            filepath = os.path.join(path, '\CIMCON Validity Reliability BinaryclassTest' + formatted_datetime)

            # Save output data with images
            self._save_data_with_images(output_data1, [fig1], filepath, model.__class__.__name__)
            self._save_data_with_images(output_data2, [fig2], filepath, model.__class__.__name__)
            self._save_data_with_images(output_data3, [fig3], filepath, model.__class__.__name__)
        except Exception as e:
            # Print error message
            print(f'Error: {e}')